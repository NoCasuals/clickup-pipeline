import re
import io
import time
import random
import numpy as np
import pandas as pd
import openpyxl
from datetime import datetime, timedelta
import gspread
from google.oauth2.service_account import Credentials
from googleapiclient.discovery import build

# --- DATABASE AND ROOT CONFIGURATION ---
GOOGLE_SHEET_URL = "https://docs.google.com/spreadsheets/d/1RE039NcnPeQtQrvI5zjLyADzAr-ZseBPUq388SxkV-Y/edit?usp=sharing"
DRIVE_FOLDER_ID = "1zHACpi08NE9D9tg5HTb_jbkjV6RpKI2v"
DIFFICULTY_LOOKUP_URL = "https://docs.google.com/spreadsheets/d/1xJwiD1F_p6BFm4-sjEEZ0U1xCMqqUEU6TwQSUmQxW5s/edit"

def get_google_credentials():
    """
    Initializes authorized Google Service Account credentials.
    
    Returns:
        Credentials: Authenticated Google API service account token instance.
    """
    scopes = [
        "https://www.googleapis.com/auth/spreadsheets",
        "https://www.googleapis.com/auth/drive"
    ]
    return Credentials.from_service_account_file("creds.json", scopes=scopes)

def execute_with_retry(func, *args, **kwargs):
    """
    Executes a Google API method using exponential backoff to mitigate
    rate limit containment rules (HTTP 429).
    """
    max_retries = 5
    base_delay = 5
    for attempt in range(max_retries):
        try:
            return func(*args, **kwargs)
        except gspread.exceptions.APIError as e:
            if "429" in str(e) and attempt < max_retries - 1:
                delay = base_delay * (2 ** attempt) + random.uniform(0, 1)
                print(f" [!] Rate limit hit (429). Server busy. Sleeping for {delay:.2f}s before retry...")
                time.sleep(delay)
            else:
                raise e
        except Exception as e:
            raise e

def batch_delete_rows(sheet, row_nums):
    """Deletes multiple rows from a worksheet in a single batchUpdate API request.
    row_nums is a list/set of 1-based row numbers. Rows are deleted descending
    so earlier deletions don't shift the indices of later ones."""
    if not row_nums:
        return
    requests = [
        {
            "deleteDimension": {
                "range": {
                    "sheetId": sheet.id,
                    "dimension": "ROWS",
                    "startIndex": rn - 1,
                    "endIndex": rn
                }
            }
        }
        for rn in sorted(row_nums, reverse=True)
    ]
    execute_with_retry(sheet.spreadsheet.batch_update, {"requests": requests})

def load_difficulty_tiers(client, url):
    """
    Downloads structural scale tier limits from the central configuration layout.
    Processes rows sequentially as unique lower-bound threshold metrics.
    """
    try:
        print(f" -> Downloading dynamic tier maps from: {url}")
        sheet = client.open_by_url(url).sheet1
        raw_rows = sheet.get_all_values()[1:]  
        
        parsed_rows = []
        for row in raw_rows:
            if not row or not row[0].strip():
                continue
            sqft_threshold = float(row[0].strip().replace(',', ''))
            difficulty_val = float(row[1].strip())
            parsed_rows.append((sqft_threshold, difficulty_val))
            
        parsed_rows.sort(key=lambda x: x[0])
        
        tiers = []
        for i in range(len(parsed_rows)):
            min_val = parsed_rows[i][0]
            max_val = parsed_rows[i+1][0] if i + 1 < len(parsed_rows) else float('inf')
            difficulty_val = parsed_rows[i][1]
            
            tiers.append({"min": min_val, "max": max_val, "base_difficulty": difficulty_val})
            
        print(f" -> Successfully synthesized {len(tiers)} operational difficulty metrics.")
        return tiers
        
    except Exception as e:
        print(f" [!] Warning: Failed to read online scale rules. Falling back to code baseline standards. Error: {e}")
        return [
            {"min": 0, "max": 2500, "base_difficulty": 1.0},
            {"min": 2500, "max": 4000, "base_difficulty": 1.5},
            {"min": 4000, "max": float('inf'), "base_difficulty": 2.0}
        ]

def list_spreadsheets_in_folder(folder_id, service):
    """
    Retrieves and catalogs workbook names along with their alphanumeric Drive IDs.
    """
    files_map = {}
    page_token = None
    
    print(f" -> Accessing Google Drive Folder: {folder_id}...")
    while True:
        query = f"'{folder_id}' in parents and trashed = false"
        response = service.files().list(
            q=query,
            spaces='drive',
            fields='nextPageToken, files(id, name)',
            pageToken=page_token,
            supportsAllDrives=True,
            includeItemsFromAllDrives=True
        ).execute()
        
        for file in response.get('files', []):
            files_map[file['name'].upper()] = file['id']
            
        page_token = response.get('nextPageToken', None)
        if not page_token:
            break
            
    print(f" -> Found {len(files_map)} schedule files inside the folder.")
    return files_map

def parse_number_sequence(seg):
    """
    Parses and expands numeric sequences and range expressions (e.g., '1 TO 5', '1, TO 5', '12-18').
    The optional comma before TO handles patterns like '(1, to 4, 6 to 8)'.
    """
    seg = seg.upper()
    numbers = set()
    ranges = re.findall(r'(\d+)\s*,?\s*(?:TO|-)\s*(\d+)', seg)
    for start, end in ranges:
        for i in range(int(start), int(end) + 1):
            numbers.add(i)
        seg = re.sub(rf'{start}\s*,?\s*(?:TO|-)\s*{end}', '', seg)
    extra_nums = re.findall(r'\d+', seg)
    for n in extra_nums:
        numbers.add(int(n))
    return sorted(list(numbers))

def parse_compound_suffixes(blob):
    """
    Deconstructs list items and continuous ranges from compound task suffix lines.
    """
    blob = blob.upper()
    suffixes = []
    range_match = re.search(r'(\d+)\s*(?:TO|-)\s*(\d+)', blob)
    if range_match:
        start, end = int(range_match.group(1)), int(range_match.group(2))
        for i in range(start, end + 1):
            suffixes.append(str(i))
        blob = re.sub(r'\d+\s*(?:TO|-)\s*\d+', '', blob)
    parts = [p.strip() for p in blob.split(',') if p.strip()]
    for p in parts:
        if p:
            suffixes.append(p)
    return suffixes

def is_wildcard_numeric_match(allowed_model, norm_model_cell):
    """
    Validates singular vs plural structural suffix combinations (e.g., matching trailing 's' tokens).
    """
    if allowed_model.endswith('S'):
        base = allowed_model[:-1]
        if base.isdigit():
            if base in norm_model_cell:
                return True
            stripped = base.rstrip('0')
            if stripped and norm_model_cell.startswith(stripped):
                return True
    return False

def extract_allowed_targets(task_name):
    """
    Advanced pattern extraction engine. Tokenizes incoming task strings into structured 
    target validation sets across model names, distinct block maps, and structural lots.
    """
    project_match = re.search(r'(D\d{5,6})', task_name)
    if not project_match:
        return None, set(), set(), set()
    project_code = project_match.group(1)
    
    text = task_name.upper().replace(project_code.upper(), "")
    text = re.split(r'\b\d+(?:ST|ND|RD|TH)\s+REVIEW', text)[0].strip()
    
    allowed_models = set()
    allowed_blocks = set()
    allowed_lots = set()
    
    block_segments = re.findall(r'\b(?:BLOCK|BLK\.?)\s*([\d\s,ToTO\-]+)', text)
    for seg in block_segments:
        nums = parse_number_sequence(seg)
        for num in nums:
            allowed_blocks.add(f"BLOCK {num}")
            
    text_no_blocks = re.sub(r'\b(?:BLOCK|BLK\.?)\s*[\d\s,ToTO\-]+', '', text)
    
    lot_segments = re.findall(r'\b(?:LOT|LTS|LOTS)\s*([\d\s,ToTO\-]+)', text_no_blocks)
    for seg in lot_segments:
        nums = parse_number_sequence(seg)
        for num in nums:
            allowed_lots.add(num)

    def add_model_target(series, suffix):
        series = series.upper().strip()
        suffix = suffix.upper().strip()
        if suffix:
            allowed_models.add(f"{series}{suffix}")
            allowed_models.add(f"{series}-{suffix}")
            allowed_models.add(f"{series} {suffix}")
        else:
            allowed_models.add(series)
        
        if len(series) > 1 and series.endswith('S'):
            base_series = series[:-1]
            if suffix:
                allowed_models.add(f"{base_series}{suffix}")
                allowed_models.add(f"{base_series}-{suffix}")
                allowed_models.add(f"{base_series} {suffix}")
            else:
                allowed_models.add(base_series)

    range_matches = re.findall(r'\b([A-Z-]{1,4})\s*-?\s*(\d+)\s*(?:TO|-)\s*(?:[A-Z-]{1,4}\s*-?\s*)?(\d+)\b', text_no_blocks)
    for series, start_str, end_str in range_matches:
        start, end = int(start_str), int(end_str)
        padding = len(start_str)
        for i in range(start, end + 1):
            num_str = f"{i:0{padding}d}"
            add_model_target(series, num_str)
            if start_str.startswith('0'):
                add_model_target(series, str(i))

    for pfx_m in re.finditer(r'\b(\d{2,4})-(\d+)', text_no_blocks):
        base      = pfx_m.group(1)
        first_num = pfx_m.group(2)
        numbers   = [first_num]
        pos       = pfx_m.end()

        while pos < len(text_no_blocks):
            to_m = re.match(r'\s+(?:TO|to)\s+(\d+)', text_no_blocks[pos:], re.IGNORECASE)
            if to_m:
                range_start = int(numbers.pop())
                range_end   = int(to_m.group(1))
                for i in range(range_start, range_end + 1):
                    numbers.append(str(i))
                pos += to_m.end()
                continue

            csv_m = re.match(r'\s*,\s*(\d+)', text_no_blocks[pos:])
            if csv_m:
                numbers.append(csv_m.group(1))
                pos += csv_m.end()
                continue

            break  

        for num_str in numbers:
            add_model_target(base, num_str)
            if not num_str.startswith('0') and len(num_str) < 2:
                add_model_target(base, num_str.zfill(2))

    pure_num_ranges = re.findall(r'\b(\d+)\s*(?:TO|-)\s*(\d+)\b', text_no_blocks)
    for start_str, end_str in pure_num_ranges:
        start, end = int(start_str), int(end_str)
        padding = len(start_str)
        for i in range(start, end + 1):
            allowed_models.add(f"{i:0{padding}d}")

    compound_match = re.search(r'\b([A-Z-]{1,4})\s*[-]?\s*(\d+)\s*-\s*([^-\n]+)', text_no_blocks)
    if compound_match:
        prefix_series = compound_match.group(1)
        base_num      = compound_match.group(2)
        suffixes      = parse_compound_suffixes(compound_match.group(3))
        for suf in suffixes:
            add_model_target(prefix_series, f"{base_num}-{suf}")

    p23 = re.search(r'\b([A-Z-]{1,4})\s*\(\s*(\d+)\s*,?\s*(?:TO|-)\s*(\d+)\s*\)', text_no_blocks)
    if p23:
        series = p23.group(1)
        for i in range(int(p23.group(2)), int(p23.group(3)) + 1):
            add_model_target(series, str(i))

    # General parenthesized handler: catches multi-range groups like "BBs (1, to 4, 6 to 8)"
    # that p23 (single-range only) cannot handle. Harmlessly overlaps with p23 matches since
    # add_model_target writes to a set.
    for paren_m in re.finditer(r'\b([A-Z-]{1,4})\s*\(\s*([^)]+)\s*\)', text_no_blocks):
        series = paren_m.group(1)
        nums = parse_number_sequence(paren_m.group(2))
        for num in nums:
            add_model_target(series, str(num))

    p1 = re.search(r'\b([A-Z-]{1,4})\s*(\d+(?:\s*,\s*\d+)+)', text_no_blocks)
    if p1:
        series = p1.group(1)
        for n in p1.group(2).split(','):
            add_model_target(series, n.strip())

    all_tokens = re.findall(r'\b([A-Z0-9-]+)\b', text_no_blocks)
    for token in all_tokens:
        allowed_models.add(token)
        allowed_models.add(token.replace('-', ''))
        if token.endswith('S') and len(token) > 1:
            allowed_models.add(token[:-1])
            allowed_models.add(token[:-1].replace('-', ''))

    _non_model_words = {
        "MODEL", "MODELS", "PRIORITY", "REVIEW", "FIRST", "SECOND", "THIRD",
        "AND", "THE", "FOR", "WITH", "TO", "OR", "OF", "A", "AN", "IN",
        "LOT", "LOTS", "BLOCK", "BLOCKS",
    }
    allowed_models -= _non_model_words
    
    return project_code, allowed_models, allowed_blocks, allowed_lots

def parse_elevation_count(elevation_cell):
    """
    Parses structural variations to calculate elevation complexity metrics.
    """
    if not elevation_cell or str(elevation_cell).strip() in ["-", ""]:
        return 1
    el_str = str(elevation_cell).strip().upper()
    if any(char in el_str for char in ['/', ',', '+', '&']):
        elements = re.split(r'[/,&+]', el_str)
        return len([e for e in elements if e.strip()])
    if el_str.isalpha():
        return len(el_str)
    return 1

def parse_sq_ft_and_difficulty(sq_ft_val, num_elevations, tiers):
    """
    Maps the square footage value to the dynamically loaded difficulty tiers
    and computes the total difficulty calculation factoring in elevation bonuses.
    """
    s = str(sq_ft_val).upper().replace(",", "").strip()
    s_match = re.search(r'([\d/]+)', s)
    if not s_match:
        return "0", 1.0
        
    val_part = s_match.group(1)
    if "/" in val_part:
        parts = [float(x.strip()) for x in val_part.split("/") if x.strip()]
        sq_ft = max(parts) if parts else 0
    else:
        sq_ft = float(val_part)
        
    base_difficulty = 1.0
    for tier in tiers:
        if tier["min"] <= sq_ft < tier["max"]:
            base_difficulty = tier["base_difficulty"]
            break
        
    elevation_bonus = max(0, (num_elevations - 1) * 0.1)
    difficulty = round(base_difficulty + elevation_bonus, 2)
    
    sq_ft_str = str(sq_ft)
    if sq_ft_str.endswith(".0"):
        sq_ft_str = sq_ft_str[:-2]
        
    return sq_ft_str, difficulty

def find_date_column_indices(schedule_data):
    """
    Dynamically maps configuration offsets for milestones across structural schedules.
    """
    arch_idx, floor_idx, truss_idx = None, None, None
    for row in schedule_data[:5]:
        row_upper = [str(cell).upper().strip() for cell in row]
        if any("ARCH DWG RECEIVED" in c or "MODEL" in c for c in row_upper):
            for idx, text in enumerate(row_upper):
                if "ARCH DWG RECEIVED" in text:
                    arch_idx = idx
                elif "FLOOR LAYOUT RECEIVED" in text or "FLOOR RECEIVED" in text:
                    floor_idx = idx
                elif "TRUSS LAYOUT RECEIVED" in text or "TRUSS RECEIVED" in text:
                    truss_idx = idx
            break
    return arch_idx, floor_idx, truss_idx

def clean_and_parse_date(date_str):
    """
    Parses varying date format strings into standardized ISO objects.
    """
    if not date_str or date_str.strip() in ["", "NAN", "-", "NONE"]:
        return None
    date_clean = date_str.split(" ")[0].strip().replace("/", "-")
    if re.match(r'^\d{4}-\d{2}-\d{2}', date_clean):
        try:
            return datetime.strptime(date_clean[:10], "%Y-%m-%d").date()
        except ValueError:
            return None
    return None

def load_and_parse_schedule(file_id, drive_service):
    """Downloads a schedule workbook and returns (schedule_data, arch_idx, floor_idx, truss_idx)."""
    file_bytes = drive_service.files().get_media(fileId=file_id).execute()
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.worksheets[0]
    for merged_range in list(ws.merged_cells.ranges):
        min_row, max_row = merged_range.min_row, merged_range.max_row
        min_col, max_col = merged_range.min_col, merged_range.max_col
        top_left_value = ws.cell(row=min_row, column=min_col).value
        ws.unmerge_cells(start_row=min_row, start_column=min_col, end_row=max_row, end_column=max_col)
        for r in range(min_row, max_row + 1):
            for c in range(min_col, max_col + 1):
                ws.cell(row=r, column=c).value = top_left_value
    schedule_data = []
    for sheet_row in ws.iter_rows(values_only=True):
        clean_row = [str(cell).strip() if cell is not None else "" for cell in sheet_row]
        schedule_data.append(clean_row)
    arch_idx, floor_idx, truss_idx = find_date_column_indices(schedule_data)
    return schedule_data, arch_idx, floor_idx, truss_idx

def compute_matched_model_names(task_name, schedule_data):
    """
    Returns the set of model output names (uppercase) that a parent task row matches
    against the given schedule data. Used during pre-scan conflict resolution.
    """
    project_code, allowed_models, allowed_blocks, allowed_lots = extract_allowed_targets(task_name)
    if not project_code:
        return set()
    matched = set()
    for s_row_idx, s_row in enumerate(schedule_data):
        if len(s_row) < 5 or s_row_idx == 1:
            continue
        model_cell = s_row[2].strip().replace(".0", "")
        normalized_model_check = re.sub(r'[^A-Z]', '', model_cell.upper())
        if normalized_model_check in ("MODEL", "MODELNAME", "SERIES", "DESCRIPTION") or not model_cell:
            continue
        model_upper = model_cell.upper()
        norm_model_cell = re.sub(r'[^A-Z0-9]', '', model_upper)
        is_match = False
        matched_via_elevation_suffix = False
        matched_am_token = ""
        for am in allowed_models:
            am_upper = am.upper()
            match_suff = re.search(r'^(.+?)(([A-Z])+)$', am_upper)
            if match_suff:
                base_am = match_suff.group(1)
                suff_str = match_suff.group(2)
                if norm_model_cell == re.sub(r'[^A-Z0-9]', '', base_am):
                    elev_cell = s_row[3].upper() if len(s_row) > 3 else ""
                    if all(c in elev_cell for c in suff_str):
                        is_match = True
                        matched_via_elevation_suffix = True
                        matched_am_token = am
                        break
            if norm_model_cell == re.sub(r'[^A-Z0-9]', '', am_upper):
                is_match = True
                break
            if is_wildcard_numeric_match(am_upper, norm_model_cell):
                is_match = True
                break
        if not is_match:
            for ab in allowed_blocks:
                if ab in model_upper or re.sub(r'[^A-Z0-9]', '', ab) in norm_model_cell:
                    is_match = True
                    break
        if not is_match:
            for al_num in allowed_lots:
                if f"LOT {al_num}" in model_upper or f"LOT0{al_num}" in model_upper or f"LOT {al_num})" in model_upper:
                    is_match = True
                    break
        if is_match:
            model_output = matched_am_token if matched_via_elevation_suffix else model_cell
            matched.add(model_output.upper())
    return matched

def delete_model_history_from_sheets(model_full_name, proj_sheet, flat_sheet):
    """
    Purges all Sheet 2 and Sheet 3 rows for a model that has been reassigned to a
    different parent task or removed from a parent task's scope.
    """
    model_upper = model_full_name.strip().upper()
    print(f"     [x] Purging historical entries for: {model_full_name}")
    for sheet, label in [(flat_sheet, "Sheet 3"), (proj_sheet, "Sheet 2")]:
        rows = sheet.get_all_values()
        to_delete = [i + 1 for i, r in enumerate(rows) if r and r[0].strip().upper() == model_upper]
        if to_delete:
            print(f"       -> Removing {len(to_delete)} rows from {label}")
            for row_num in sorted(to_delete, reverse=True):
                execute_with_retry(sheet.delete_rows, row_num)

def batch_purge_models_from_sheets(model_names, proj_sheet, flat_sheet):
    """
    Purges Sheet 2 and Sheet 3 rows for a collection of models in exactly two reads
    and two writes (one per sheet), regardless of how many models need purging.
    """
    if not model_names:
        return
    model_uppers = {m.strip().upper() for m in model_names}
    print(f"     [x] Batch purging {len(model_uppers)} model(s) from Sheet 2/3...")
    for sheet, label in [(flat_sheet, "Sheet 3"), (proj_sheet, "Sheet 2")]:
        rows = sheet.get_all_values()
        to_delete = [i + 1 for i, r in enumerate(rows) if r and r[0].strip().upper() in model_uppers]
        if to_delete:
            print(f"       -> Removing {len(to_delete)} rows from {label}")
            batch_delete_rows(sheet, to_delete)

def main():
    print(f"[{datetime.now()}] Initializing Advanced Inline Upsert Engine...")
    creds = get_google_credentials()
    client = gspread.authorize(creds)
    drive_service = build('drive', 'v3', credentials=creds)
    
    try:
        difficulty_tiers = load_difficulty_tiers(client, DIFFICULTY_LOOKUP_URL)
        drive_files = list_spreadsheets_in_folder(DRIVE_FOLDER_ID, drive_service)
        master_workbook = client.open_by_url(GOOGLE_SHEET_URL)
        log_sheet = master_workbook.sheet1  
        
        # --- AUTOMATED DATA PURGE PROTOCOL (SHEET 1) ---
        print(" -> Scanning Sheet 1 for legacy header row duplicates...")
        all_log_rows = log_sheet.get_all_values()
        sheet1_purge_indices = []
        for r_idx, r_data in enumerate(all_log_rows):
            if len(r_data) > 2:
                model_col_text = re.sub(r'[^A-Z]', '', r_data[2].strip().upper())
                if model_col_text == "MODEL":
                    sheet1_purge_indices.append(r_idx + 1)
        
        if sheet1_purge_indices:
            print(f" [!] Purging {len(sheet1_purge_indices)} incorrect structural rows from Sheet 1...")
            batch_delete_rows(log_sheet, sheet1_purge_indices)
            all_log_rows = log_sheet.get_all_values()

        # --- AUTOMATED DATA PURGE PROTOCOL (SHEET 2 - WEIGHTED) ---
        print(" -> Scanning Sheet 2 (Consolidated Log) for corrupt entries...")
        proj_sheet = master_workbook.worksheets()[1]
        proj_rows = proj_sheet.get_all_values()
        sheet2_purge_indices = []
        for r_idx, r_data in enumerate(proj_rows):
            if len(r_data) > 0:
                normalized_name = r_data[0].strip().upper()
                if normalized_name.endswith(" - MODEL") or normalized_name == "MODEL":
                    sheet2_purge_indices.append(r_idx + 1)
                    
        if sheet2_purge_indices:
            print(f" [!] Purging {len(sheet2_purge_indices)} corrupt historical series items from Sheet 2...")
            batch_delete_rows(proj_sheet, sheet2_purge_indices)
            proj_rows = proj_sheet.get_all_values()

        # --- UPDATED: AUTOMATED DATA PURGE PROTOCOL (SHEET 3 - UNWEIGHTED FLAT) ---
        print(" -> Scanning Sheet 3 (Flat Log) for corrupt entries...")
        flat_sheet = master_workbook.worksheets()[2]
        flat_rows = flat_sheet.get_all_values()
        sheet3_purge_indices = []
        for r_idx, r_data in enumerate(flat_rows):
            if len(r_data) > 0:
                normalized_name = r_data[0].strip().upper()
                if normalized_name.endswith(" - MODEL") or normalized_name == "MODEL":
                    sheet3_purge_indices.append(r_idx + 1)
                    
        if sheet3_purge_indices:
            print(f" [!] Purging {len(sheet3_purge_indices)} corrupt historical series items from Sheet 3...")
            batch_delete_rows(flat_sheet, sheet3_purge_indices)
            flat_rows = flat_sheet.get_all_values()

        # --- PRE-SCAN: RESOLVE MODEL OWNERSHIP FOR SHARED PROJECT CODES ---
        print(" -> Pre-scanning for shared project codes across parent tasks...")
        project_code_rows = {}  # project_code -> list of [row_idx, clickup_start_str, file_id, row]
        for _idx, _row in enumerate(all_log_rows):
            if _idx == 0 or len(_row) < 2 or not _row[1].strip():
                continue
            _pc, _, _, _ = extract_allowed_targets(_row[1])
            if not _pc:
                continue
            _fid = next((fid for fname, fid in drive_files.items() if _pc in fname), None)
            if not _fid:
                continue
            _start = _row[10].strip() if len(_row) > 10 else ""
            project_code_rows.setdefault(_pc, []).append([_idx, _start, _fid, _row])

        schedule_data_cache = {}   # file_id -> (schedule_data, arch_idx, floor_idx, truss_idx)
        model_blacklists = {}      # row_idx -> set of uppercase full model names to skip
        pre_purge_models = set()   # full model names needing Sheet 2/3 cleanup before re-assignment

        conflicted_projects = {pc for pc, entries in project_code_rows.items() if len(entries) > 1}
        if conflicted_projects:
            print(f" -> Found {len(conflicted_projects)} project code(s) with multiple parent tasks. Resolving model ownership...")
            for _pc in conflicted_projects:
                _entries = project_code_rows[_pc]
                _entries.sort(key=lambda e: (clean_and_parse_date(e[1]) or datetime.min.date()), reverse=True)
                _fid = _entries[0][2]
                if _fid not in schedule_data_cache:
                    try:
                        schedule_data_cache[_fid] = load_and_parse_schedule(_fid, drive_service)
                    except Exception as _e:
                        print(f"     [!] Could not load schedule for {_pc}: {_e}")
                        continue
                _sched, _, _, _ = schedule_data_cache[_fid]
                _claimed_globally = set()
                for _entry in _entries:
                    _row_idx, _, _, _row_data = _entry
                    _this_models = {f"{_pc} - {m}".upper() for m in compute_matched_model_names(_row_data[1], _sched)}
                    _contested = _this_models & _claimed_globally
                    if _contested:
                        model_blacklists.setdefault(_row_idx, set()).update(_contested)
                        pre_purge_models.update(_contested)
                    _claimed_globally.update(_this_models - _contested)

        if pre_purge_models:
            print(f" -> Purging {len(pre_purge_models)} contested model(s) from Sheet 2/3 before re-assignment...")
            batch_purge_models_from_sheets(pre_purge_models, proj_sheet, flat_sheet)
            # Refresh after row deletions so indices start accurate
            proj_rows = proj_sheet.get_all_values()
            flat_rows = flat_sheet.get_all_values()

        # Build Sheet 2/3 in-memory indices once. Maintained across all iterations to
        # avoid a get_all_values() call per parent task (the main source of 429 read errors).
        existing_index = {}
        for r_idx, r_cells in enumerate(proj_rows):
            if r_idx == 0: continue
            nm = r_cells[0].strip().upper() if len(r_cells) > 0 else ""
            dt = r_cells[1].strip() if len(r_cells) > 1 else ""
            kp = r_cells[2].strip() if len(r_cells) > 2 else ""
            if nm and dt: existing_index[(nm, dt)] = (r_idx + 1, kp)

        flat_existing_index = {}
        for r_idx, r_cells in enumerate(flat_rows):
            if r_idx == 0: continue
            nm = r_cells[0].strip().upper() if len(r_cells) > 0 else ""
            dt = r_cells[1].strip() if len(r_cells) > 1 else ""
            kp = r_cells[2].strip() if len(r_cells) > 2 else ""
            if nm and dt: flat_existing_index[(nm, dt)] = (r_idx + 1, kp)

        # Dirty flags: set True after row deletions (which shift indices). False = index is current.
        sheet2_dirty = False
        sheet3_dirty = False

        print("\nScanning rows and executing architectural upsert calculations...")
        for idx in range(len(all_log_rows) - 1, 0, -1):
            row = all_log_rows[idx]
            if len(row) < 2 or not row[1].strip():
                continue
                
            task_name = row[1]  
            project_code, allowed_models, allowed_blocks, allowed_lots = extract_allowed_targets(task_name)
            
            if not project_code or (not allowed_models and not allowed_blocks and not allowed_lots):
                continue
                
            matched_file_id = None
            for filename, file_id in drive_files.items():
                if project_code in filename:
                    matched_file_id = file_id
                    break
                    
            if not matched_file_id:
                continue
                
            # existing_children: cleaned_key -> (sheet1_row_num, original_model_name)
            existing_children = {}
            scan_idx = idx + 1
            while scan_idx < len(all_log_rows):
                scan_row = all_log_rows[scan_idx]
                if (len(scan_row) > 0 and scan_row[0].strip()) or (len(scan_row) > 1 and scan_row[1].strip()):
                    break
                if len(scan_row) > 2 and scan_row[2].strip():
                    raw_model_name = scan_row[2].strip().upper()
                    cleaned_key = re.sub(r'[^A-Z0-9]', '', raw_model_name)
                    existing_children[cleaned_key] = (scan_idx + 1, raw_model_name)
                scan_idx += 1

            try:
                if matched_file_id in schedule_data_cache:
                    schedule_data, arch_idx, floor_idx, truss_idx = schedule_data_cache[matched_file_id]
                else:
                    schedule_data, arch_idx, floor_idx, truss_idx = load_and_parse_schedule(matched_file_id, drive_service)
                    schedule_data_cache[matched_file_id] = (schedule_data, arch_idx, floor_idx, truss_idx)

                subtask_compiled_matches = []
                
                for s_row_idx, s_row in enumerate(schedule_data):
                    if len(s_row) < 5 or s_row_idx == 1:
                        continue

                    series_cell = s_row[1].strip()
                    model_cell = s_row[2].strip().replace(".0", "")

                    normalized_model_check = re.sub(r'[^A-Z]', '', model_cell.upper())
                    if normalized_model_check in ("MODEL", "MODELNAME", "SERIES", "DESCRIPTION") or not model_cell:
                        continue
                        
                    if series_cell:
                        current_series = series_cell
                        
                    model_upper = model_cell.upper()
                    norm_model_cell = re.sub(r'[^A-Z0-9]', '', model_upper)
                    
                    is_match = False
                    is_block = False
                    matched_via_elevation_suffix = False
                    suffix_elevation_count = 1
                    matched_am_token = ""
                    
                    for am in allowed_models:
                        am_upper = am.upper()
                        match_suff = re.search(r'^(.+?)(([A-Z])+)$', am_upper)
                        if match_suff:
                            base_am = match_suff.group(1)
                            suff_str = match_suff.group(2)
                            if norm_model_cell == re.sub(r'[^A-Z0-9]', '', base_am):
                                if all(c in s_row[3].upper() for c in suff_str):
                                    is_match = True
                                    matched_via_elevation_suffix = True
                                    suffix_elevation_count = len(suff_str)
                                    matched_am_token = am
                                    break
                                    
                        if norm_model_cell == re.sub(r'[^A-Z0-9]', '', am_upper):
                            is_match = True
                            break
                        if is_wildcard_numeric_match(am_upper, norm_model_cell):
                            is_match = True
                            break
                                    
                    if not is_match:
                        for ab in allowed_blocks:
                            if ab in model_upper or re.sub(r'[^A-Z0-9]', '', ab) in norm_model_cell:
                                is_match = True
                                is_block = True
                                break
                                
                    if not is_match:
                        for al_num in allowed_lots:
                            if f"LOT {al_num}" in model_upper or f"LOT0{al_num}" in model_upper or f"LOT {al_num})" in model_upper:
                                is_match = True
                                break
                                
                    if is_match:
                        if matched_via_elevation_suffix:
                            model_output = matched_am_token
                            num_elevations = suffix_elevation_count
                        else:
                            model_output = model_cell
                            num_elevations = parse_elevation_count(s_row[3])
                        
                        if is_block:
                            sq_ft_output = ""
                            difficulty_output = "1"
                        else:
                            sq_ft_output, difficulty_output = parse_sq_ft_and_difficulty(s_row[4], num_elevations, difficulty_tiers)
                            
                        arch_cell_str = s_row[arch_idx].strip() if arch_idx is not None and arch_idx < len(s_row) else ""
                        floor_cell_str = s_row[floor_idx].strip() if floor_idx is not None and floor_idx < len(s_row) else ""
                        truss_cell_str = s_row[truss_idx].strip() if truss_idx is not None and truss_idx < len(s_row) else ""
                        
                        subtask_compiled_matches.append({
                            "model_output": model_output,
                            "is_block": is_block,
                            "num_elevations": num_elevations,
                            "sq_ft_output": sq_ft_output,
                            "difficulty_output": difficulty_output,
                            "arch_dt": clean_and_parse_date(arch_cell_str),
                            "floor_dt": clean_and_parse_date(floor_cell_str),
                            "truss_dt": clean_and_parse_date(truss_cell_str),
                            "final_start_date": ""
                        })

                # --- SCOPE CHANGE DETECTION ---
                # Find child rows that existed previously but are no longer matched.
                # Purge their Sheet 1 child rows and their Sheet 2/3 history.
                new_matched_keys = {re.sub(r'[^A-Z0-9]', '', item['model_output'].upper())
                                    for item in subtask_compiled_matches}
                scope_removed_names = []
                scope_removed_sh1_rows = []
                for ck, (row_num, orig_name) in existing_children.items():
                    if ck not in new_matched_keys:
                        scope_removed_names.append(f"{project_code} - {orig_name}")
                        scope_removed_sh1_rows.append(row_num)

                if scope_removed_names:
                    batch_purge_models_from_sheets(scope_removed_names, proj_sheet, flat_sheet)
                    sheet2_dirty = True
                    sheet3_dirty = True
                    batch_delete_rows(log_sheet, scope_removed_sh1_rows)

                valid_model_dates = []
                for item in subtask_compiled_matches:
                    if not item["is_block"]:
                        if item["arch_dt"] and item["floor_dt"] and item["truss_dt"]:
                            calc_date = max(item["arch_dt"], item["floor_dt"], item["truss_dt"]) + timedelta(days=1)
                            item["final_start_date"] = calc_date.strftime("%Y-%m-%d")
                            valid_model_dates.append(calc_date)

                max_subtask_model_date = max(valid_model_dates) if valid_model_dates else None

                for item in subtask_compiled_matches:
                    if item["is_block"]:
                        if item["arch_dt"] and item["truss_dt"]:
                            block_base_date = max(item["arch_dt"], item["truss_dt"]) + timedelta(days=1)
                            final_block_date = max_subtask_model_date if max_subtask_model_date and max_subtask_model_date > block_base_date else block_base_date
                            item["final_start_date"] = final_block_date.strftime("%Y-%m-%d")

                new_rows_to_insert = []
                project_sheet_updates = []
                seen_models_this_pass = set()
                row_blacklist = model_blacklists.get(idx, set())

                # Remove stale Sheet 1 child rows for models now owned by a later-dated parent.
                # These were skipped in the scope-change pass (they ARE in the schedule) but
                # must not remain under this losing parent.
                if row_blacklist:
                    blacklisted_sh1_rows = [
                        row_num for ck, (row_num, orig_name) in existing_children.items()
                        if f"{project_code} - {orig_name}".upper() in row_blacklist
                    ]
                    if blacklisted_sh1_rows:
                        print(f"     [x] Removing {len(blacklisted_sh1_rows)} blacklisted child row(s) from Sheet 1...")
                        batch_delete_rows(log_sheet, blacklisted_sh1_rows)
                        blacklisted_row_set = set(blacklisted_sh1_rows)
                        existing_children = {ck: v for ck, v in existing_children.items()
                                             if v[0] not in blacklisted_row_set}

                for item in subtask_compiled_matches:
                    model_output = item["model_output"]

                    # Skip models that belong to a later-dated parent with the same project code
                    if f"{project_code} - {model_output}".upper() in row_blacklist:
                        continue

                    difficulty_output = item["difficulty_output"]
                    start_date_output = item["final_start_date"]

                    try:
                        parent_col_g = float(row[6].strip()) if len(row) > 6 and row[6].strip() else 0
                        model_difficulty = float(difficulty_output) if difficulty_output else 0
                        loading_quotient = round(model_difficulty / parent_col_g, 4) if parent_col_g != 0 else ""

                        # UPDATED: Computes unweighted loading metric allocation payload (1 / duration)
                        flat_loading_quotient = round(1.0 / parent_col_g, 4) if parent_col_g != 0 else ""
                    except Exception:
                        loading_quotient = ""
                        flat_loading_quotient = ""

                    parent_col_k_str = row[10].strip() if len(row) > 10 else ""

                    if model_output.upper() not in seen_models_this_pass:
                        seen_models_this_pass.add(model_output.upper())
                        project_sheet_updates.append({
                            "model_name": f"{project_code} - {model_output}",
                            "loading_quotient": loading_quotient,
                            "flat_loading_quotient": flat_loading_quotient,
                            "clickup_start_date": parent_col_k_str,
                            "duration": row[6].strip() if len(row) > 6 else "",
                            "clickup_due_date": row[11].strip() if len(row) > 11 else ""
                        })

                    model_lookup_key = re.sub(r'[^A-Z0-9]', '', model_output.upper())
                    if model_lookup_key in existing_children:
                        live_row_num, _ = existing_children[model_lookup_key]
                        orig_child_row = all_log_rows[live_row_num - 1]
                        
                        current_el = orig_child_row[3].strip() if len(orig_child_row) > 3 else ""
                        current_sq = orig_child_row[4].strip() if len(orig_child_row) > 4 else ""
                        current_diff = orig_child_row[5].strip() if len(orig_child_row) > 5 else ""
                        current_h = orig_child_row[7].strip() if len(orig_child_row) > 7 else ""     
                        current_j = orig_child_row[8].strip() if len(orig_child_row) > 8 else ""     
                        current_k = orig_child_row[10].strip() if len(orig_child_row) > 10 else ""   
                        
                        if (current_el != str(item["num_elevations"]) or current_sq != str(item["sq_ft_output"]) or 
                            current_diff != str(difficulty_output) or current_h != str(loading_quotient) or 
                            current_j != str(start_date_output) or current_k != str(parent_col_k_str)):
                            
                            update_range = f"D{live_row_num}:K{live_row_num}"
                            update_payload = [[str(item["num_elevations"]), str(item["sq_ft_output"]), str(difficulty_output), "", str(loading_quotient), "", str(start_date_output), str(parent_col_k_str)]]
                            execute_with_retry(log_sheet.update, range_name=update_range, values=update_payload, value_input_option="USER_ENTERED")
                    else:
                        new_rows_to_insert.append(["", "", model_output, item["num_elevations"], item["sq_ft_output"], difficulty_output, "", loading_quotient, "", start_date_output, parent_col_k_str])
                
                if new_rows_to_insert:
                    execute_with_retry(log_sheet.insert_rows, new_rows_to_insert, row=idx + 2, value_input_option="USER_ENTERED")
                
                # --- DATA ROUTING TRACKING SWEEP FOR SHEET 2 & SHEET 3 ---
                if project_sheet_updates:
                    # Re-read only when rows were deleted (dirty flag). Inserts are handled
                    # by updating the in-memory index directly, so no re-read is needed for those.
                    if sheet2_dirty:
                        proj_rows = proj_sheet.get_all_values()
                        existing_index = {}
                        for r_idx, r_cells in enumerate(proj_rows):
                            if r_idx == 0: continue
                            nm = r_cells[0].strip().upper() if len(r_cells) > 0 else ""
                            dt = r_cells[1].strip() if len(r_cells) > 1 else ""
                            kp = r_cells[2].strip() if len(r_cells) > 2 else ""
                            if nm and dt: existing_index[(nm, dt)] = (r_idx + 1, kp)
                        sheet2_dirty = False

                    if sheet3_dirty:
                        flat_rows = flat_sheet.get_all_values()
                        flat_existing_index = {}
                        for r_idx, r_cells in enumerate(flat_rows):
                            if r_idx == 0: continue
                            nm = r_cells[0].strip().upper() if len(r_cells) > 0 else ""
                            dt = r_cells[1].strip() if len(r_cells) > 1 else ""
                            kp = r_cells[2].strip() if len(r_cells) > 2 else ""
                            if nm and dt: flat_existing_index[(nm, dt)] = (r_idx + 1, kp)
                        sheet3_dirty = False

                    today_str = datetime.now().strftime("%Y-%m-%d")
                    pending_updates = []
                    rows_to_insert_at_top = []
                    flat_pending_updates = []
                    flat_rows_to_insert_at_top = []

                    for p_update in project_sheet_updates:
                        full_name        = p_update["model_name"]
                        lq_val           = str(p_update["loading_quotient"])
                        flat_lq_val      = str(p_update["flat_loading_quotient"])
                        clickup_start_dt = clean_and_parse_date(p_update.get("clickup_start_date", ""))
                        clickup_due_dt   = clean_and_parse_date(p_update.get("clickup_due_date", ""))

                        # Sweep Sheet 2 for required weighted metric synchronizations
                        for (nm, dt), index_val in list(existing_index.items()):
                            if not isinstance(index_val, tuple) or len(index_val) < 2:
                                continue
                            row_num, current_kpi = index_val
                            if nm == full_name.upper():
                                if current_kpi != lq_val:
                                    pending_updates.append((row_num, lq_val))
                                    existing_index[(nm, dt)] = (row_num, lq_val)

                        # Sweep Sheet 3 for required unweighted flat metric synchronizations
                        for (nm, dt), index_val in list(flat_existing_index.items()):
                            if not isinstance(index_val, tuple) or len(index_val) < 2:
                                continue
                            row_num, current_kpi = index_val
                            if nm == full_name.upper():
                                if current_kpi != flat_lq_val:
                                    flat_pending_updates.append((row_num, flat_lq_val))
                                    flat_existing_index[(nm, dt)] = (row_num, flat_lq_val)

                        dates_to_ensure = set()
                        if clickup_start_dt and clickup_due_dt and clickup_due_dt >= clickup_start_dt:
                            for bd in pd.bdate_range(str(clickup_start_dt), str(clickup_due_dt)):
                                dates_to_ensure.add(str(bd.date()))
                        if not dates_to_ensure:
                            dates_to_ensure.add(today_str)

                        for day_str in sorted(dates_to_ensure):
                            key = (full_name.upper(), day_str)
                            
                            # Append logic for missing intervals on Sheet 2
                            if key not in existing_index:
                                rows_to_insert_at_top.append([full_name, day_str, lq_val])
                                existing_index[key] = (0, lq_val)
                                
                            # Append logic for missing intervals on Sheet 3
                            if key not in flat_existing_index:
                                flat_rows_to_insert_at_top.append([full_name, day_str, flat_lq_val])
                                flat_existing_index[key] = (0, flat_lq_val)

                    # --- EXECUTE SHEET 2 COMMITS (BATCHED SINGLE-CALL API PASS) ---
                    cells_to_update = []
                    for row_num, new_kpi in pending_updates:
                        cells_to_update.append(gspread.cell.Cell(row=row_num, col=3, value=new_kpi))

                    if cells_to_update:
                        execute_with_retry(proj_sheet.update_cells, cells_to_update, value_input_option="USER_ENTERED")
                        print(f" [≠] TAB 2 UPDATE: Batched synchronization complete for {len(cells_to_update)} records.")

                    if rows_to_insert_at_top:
                        n2 = len(rows_to_insert_at_top)
                        execute_with_retry(proj_sheet.insert_rows, rows_to_insert_at_top, row=2, value_input_option="USER_ENTERED")
                        # Shift all stored row numbers by n2 (rows inserted at position 2),
                        # then register the new entries so the next iteration skips re-reading.
                        existing_index = {(nm, dt): (rn + n2 if rn > 0 else 0, kp)
                                          for (nm, dt), (rn, kp) in existing_index.items()}
                        for i, ins_row in enumerate(rows_to_insert_at_top):
                            nm = ins_row[0].strip().upper() if ins_row else ""
                            dt = ins_row[1].strip() if len(ins_row) > 1 else ""
                            kp = str(ins_row[2]) if len(ins_row) > 2 else ""
                            if nm and dt:
                                existing_index[(nm, dt)] = (2 + i, kp)

                    # --- EXECUTE SHEET 3 COMMITS (BATCHED SINGLE-CALL API PASS) ---
                    flat_cells_to_update = []
                    for row_num, new_kpi in flat_pending_updates:
                        flat_cells_to_update.append(gspread.cell.Cell(row=row_num, col=3, value=new_kpi))

                    if flat_cells_to_update:
                        execute_with_retry(flat_sheet.update_cells, flat_cells_to_update, value_input_option="USER_ENTERED")
                        print(f" [≠] TAB 3 UPDATE: Batched synchronization complete for {len(flat_cells_to_update)} records.")

                    if flat_rows_to_insert_at_top:
                        n3 = len(flat_rows_to_insert_at_top)
                        execute_with_retry(flat_sheet.insert_rows, flat_rows_to_insert_at_top, row=2, value_input_option="USER_ENTERED")
                        flat_existing_index = {(nm, dt): (rn + n3 if rn > 0 else 0, kp)
                                               for (nm, dt), (rn, kp) in flat_existing_index.items()}
                        for i, ins_row in enumerate(flat_rows_to_insert_at_top):
                            nm = ins_row[0].strip().upper() if ins_row else ""
                            dt = ins_row[1].strip() if len(ins_row) > 1 else ""
                            kp = str(ins_row[2]) if len(ins_row) > 2 else ""
                            if nm and dt:
                                flat_existing_index[(nm, dt)] = (2 + i, kp)
                    
            except Exception as inner_e:
                print(f"     [!] Error processing row index {idx + 1}: {str(inner_e)}")

        print("\nSUCCESS: Multi-file structural matrix execution complete.")
    except Exception as e:
        print(f"\nEngine Failed: {str(e)}")

if __name__ == "__main__":
    main()
